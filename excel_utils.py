"""
Excelファイル解析のためのユーティリティ関数
"""

import openpyxl
from typing import Dict, Any, List
import io


def analyze_excel_structure(file_content: bytes) -> Dict[str, Any]:
    """
    Excelファイルの構造を分析し、基本的な統計情報を取得する
    
    Args:
        file_content: Excelファイルのバイナリコンテンツ
        
    Returns:
        Dict: 各シートの基本統計情報
    """
    try:
        # バイナリデータからWorkbookを読み込み
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
    except Exception as e:
        return {"error": f"ファイルの読み込みに失敗しました: {e}"}

    analysis_results = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_results = {}

        # 基本的な情報を取得
        max_row = sheet.max_row
        max_column = sheet.max_column
        sheet_results['max_row'] = max_row
        sheet_results['max_column'] = max_column

        # 非空セルの数をカウント
        non_empty_cells = 0
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
            for cell in row:
                if cell.value is not None:
                    non_empty_cells += 1
        
        total_possible_cells = max_row * max_column
        if total_possible_cells > 0:
            data_density = non_empty_cells / total_possible_cells
        else:
            data_density = 0
        sheet_results['data_density'] = data_density
        sheet_results['non_empty_cells'] = non_empty_cells

        # 結合セルの数
        num_merged_cells = len(sheet.merged_cells.ranges)
        sheet_results['num_merged_cells'] = num_merged_cells

        # Excelテーブルの有無
        if hasattr(sheet, 'tables') and sheet.tables:
            sheet_results['has_excel_tables'] = True
            sheet_results['num_excel_tables'] = len(sheet.tables)
        else:
            sheet_results['has_excel_tables'] = False

        analysis_results[sheet_name] = sheet_results

    return analysis_results


def excel_to_text_representation(file_content: bytes, max_rows: int = 20) -> Dict[str, str]:
    """
    ExcelファイルをLLM分析用のテキスト形式に変換する
    
    Args:
        file_content: Excelファイルのバイナリコンテンツ
        max_rows: 変換する最大行数（デフォルト: 20行）
        
    Returns:
        Dict: シート名をキーとするテキスト表現の辞書
    """
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
    except Exception as e:
        return {"error": f"ファイルの読み込みに失敗しました: {e}"}

    text_representations = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        text_lines = []
        
        # 結合セル情報を取得
        merged_ranges = {cell_range.min_row: cell_range for cell_range in sheet.merged_cells.ranges}
        
        # 最大行数を制限
        actual_max_row = min(sheet.max_row, max_rows)
        
        for row_idx in range(1, actual_max_row + 1):
            row_cells = []
            
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                # セルの値を取得
                if cell.value is None:
                    cell_text = "[EMPTY]"
                else:
                    cell_text = str(cell.value)
                
                # 結合セルの処理
                is_merged = False
                for merged_range in sheet.merged_cells.ranges:
                    if (merged_range.min_row <= row_idx <= merged_range.max_row and 
                        merged_range.min_col <= col_idx <= merged_range.max_col):
                        if row_idx == merged_range.min_row and col_idx == merged_range.min_col:
                            # 結合セルの左上セル
                            pass  # 通常の値を使用
                        else:
                            # 結合セルの他の部分
                            cell_text = "[MERGED]"
                        is_merged = True
                        break
                
                # 書式情報の追加（太字の検出）
                if cell.font and cell.font.bold:
                    cell_text = f"**{cell_text}**"
                
                # 背景色の検出（簡易版）
                if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
                    cell_text = f"[BG]{cell_text}"
                
                row_cells.append(cell_text)
            
            # 行のテキスト表現を作成
            row_text = f"row{row_idx}: | " + " | ".join(row_cells) + " |"
            text_lines.append(row_text)
        
        # 行数が制限されている場合は注記を追加
        if sheet.max_row > max_rows:
            text_lines.append(f"... (実際には{sheet.max_row}行ありますが、最初の{max_rows}行のみ表示)")
        
        text_representations[sheet_name] = "\n".join(text_lines)

    return text_representations


def format_analysis_results(analysis_output) -> str:
    """
    分析結果を見やすい形式でフォーマットする
    
    Args:
        analysis_output: ExcelAnalysisOutput オブジェクト
        
    Returns:
        str: フォーマットされた分析結果
    """
    if not analysis_output:
        return "分析結果がありません。"
    
    result_lines = []
    result_lines.append("## Excel分析結果\n")
    
    for sheet_result in analysis_output.sheets:
        result_lines.append(f"### シート: {sheet_result.sheet_name}")
        result_lines.append(f"**分類**: {sheet_result.sheet_type}")
        
        if sheet_result.header_info:
            result_lines.append("**ヘッダー情報**:")
            result_lines.append(f"- 開始行: {sheet_result.header_info.start_row}")
            result_lines.append(f"- 終了行: {sheet_result.header_info.end_row}")
            result_lines.append(f"- タイプ: {sheet_result.header_info.header_type}")
        else:
            result_lines.append("**ヘッダー情報**: なし")
        
        result_lines.append(f"**判定理由**: {sheet_result.reasoning}")
        result_lines.append("")  # 空行
    
    return "\n".join(result_lines)

