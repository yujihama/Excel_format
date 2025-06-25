"""
テスト用のサンプルExcelファイルを作成
"""

import openpyxl
from openpyxl.styles import Font
import os

def create_sample_excel_files():
    """テスト用のサンプルExcelファイルを作成"""
    
    # 1. 単一ヘッダーのテーブル
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "単一ヘッダーテーブル"
    
    # ヘッダー行（太字）
    headers = ["ID", "商品名", "価格", "在庫数"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # データ行
    data = [
        [1, "りんご", 100, 50],
        [2, "みかん", 80, 30],
        [3, "バナナ", 120, 25],
        [4, "ぶどう", 200, 15]
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)
    
    wb1.save("/home/ubuntu/excel_analyzer/sample_single_header.xlsx")
    
    # 2. 複合ヘッダーのテーブル
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "複合ヘッダーテーブル"
    
    # 1行目のヘッダー（結合セル）
    ws2.merge_cells('A1:A2')  # 地域
    ws2.merge_cells('B1:B2')  # 商品名
    ws2.merge_cells('C1:D1')  # 売上
    
    ws2['A1'] = "地域"
    ws2['A1'].font = Font(bold=True)
    ws2['B1'] = "商品名"
    ws2['B1'].font = Font(bold=True)
    ws2['C1'] = "売上"
    ws2['C1'].font = Font(bold=True)
    
    # 2行目のヘッダー
    ws2['C2'] = "Q1"
    ws2['C2'].font = Font(bold=True)
    ws2['D2'] = "Q2"
    ws2['D2'].font = Font(bold=True)
    
    # データ行
    data2 = [
        ["北海道", "りんご", 100, 120],
        ["東京", "みかん", 150, 130],
        ["大阪", "バナナ", 80, 90],
        ["福岡", "ぶどう", 200, 180]
    ]
    
    for row_idx, row_data in enumerate(data2, 3):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    wb2.save("/home/ubuntu/excel_analyzer/sample_multi_header.xlsx")
    
    # 3. フォーム形式
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = "フォーム"
    
    # フォームタイトル
    ws3['A1'] = "顧客情報入力フォーム"
    ws3['A1'].font = Font(bold=True, size=14)
    
    # フォーム項目
    form_items = [
        ("顧客ID:", "C001"),
        ("会社名:", "株式会社サンプル"),
        ("担当者名:", "田中太郎"),
        ("電話番号:", "03-1234-5678"),
        ("メールアドレス:", "tanaka@sample.com"),
        ("住所:", "東京都千代田区...")
    ]
    
    for idx, (label, value) in enumerate(form_items, 3):
        ws3.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=idx, column=2, value=value)
    
    wb3.save("/home/ubuntu/excel_analyzer/sample_form.xlsx")
    
    print("サンプルExcelファイルを作成しました:")
    print("- sample_single_header.xlsx (単一ヘッダーテーブル)")
    print("- sample_multi_header.xlsx (複合ヘッダーテーブル)")
    print("- sample_form.xlsx (フォーム形式)")

if __name__ == "__main__":
    create_sample_excel_files()

